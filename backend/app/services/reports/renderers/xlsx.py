"""Excel renderer.

An analyst opening the spreadsheet wants the *numbers*, not a transcription of
the prose. So this renderer inverts the emphasis: a summary sheet, then one
sheet per section carrying its tables and metrics, then a dedicated evidence
sheet. Charts become native Excel charts where the data supports it, so they
stay live when a user edits a figure.

Prose is included but subordinated — wrapped in a single wide column beneath
the data rather than paragraph by paragraph, because a spreadsheet is a poor
place to read an essay and a good place to check a calculation.
"""
from __future__ import annotations

import io
import re
import time

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.reports.blocks import (
    Block, BlockKind, ChartKind, ReportDocument, SectionKey,
)
from app.domain.reports.citations import evidence_by_source, strip_markers
from app.services.reports.renderers.base import (
    OutputFormat, RenderResult, ReportRenderer, cover_pairs, register,
)

NAVY = "1E3A8A"
SURFACE = "F5F7FA"
MUTED = "64748B"
WHITE = "FFFFFF"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SURFACE_FILL = PatternFill("solid", fgColor=SURFACE)
HEADER_FONT = Font(bold=True, color=WHITE, size=9)
LABEL_FONT = Font(bold=True, size=9)
BODY_FONT = Font(size=9)
MUTED_FONT = Font(size=8, color=MUTED)
TITLE_FONT = Font(bold=True, size=14, color=NAVY)

THIN = Side(style="thin", color="E2E8F0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

#: Excel forbids these in sheet names, and caps the length at 31.
_ILLEGAL = re.compile(r"[\[\]:*?/\\]")
#: A cell beginning with one of these is interpreted as a formula. A label such
#: as "-- see note" would become a #NAME? error, so it is prefixed.
_FORMULA_LEAD = ("=", "+", "-", "@")


def _sheet_name(title: str, used: set[str]) -> str:
    cleaned = _ILLEGAL.sub("", title)[:31].strip() or "Sheet"
    candidate = cleaned
    suffix = 2
    while candidate.lower() in used:
        candidate = f"{cleaned[:28]}_{suffix}"
        suffix += 1
    used.add(candidate.lower())
    return candidate


def _safe(value: str) -> str:
    text = str(value)
    return f"'{text}" if text[:1] in _FORMULA_LEAD else text


def _numeric(value: str) -> float | None:
    """Parse a formatted cell back to a number so Excel can compute on it.

    A report table holds "₹33,543" as text. Writing that verbatim gives a
    spreadsheet nobody can sum, which defeats the point of an Excel export.
    """
    cleaned = (
        str(value).replace(",", "").replace("₹", "").replace("%", "")
        .replace("x", "").replace("/100", "").strip()
    )
    if not cleaned or cleaned in {"—", "-", "N/A", "NA"}:
        return None
    negative = cleaned.startswith("(") and cleaned.endswith(")")
    if negative:
        cleaned = cleaned[1:-1]
    try:
        number = float(cleaned)
    except ValueError:
        return None
    return -number if negative else number


@register
class XlsxRenderer(ReportRenderer):
    """Renders a report to an Excel workbook."""

    fmt = OutputFormat.XLSX

    def render(self, document: ReportDocument) -> RenderResult:
        started = time.perf_counter()
        workbook = Workbook()
        workbook.remove(workbook.active)
        used: set[str] = set()

        self._summary_sheet(workbook, document, used)
        for section in document.ordered():
            if section.key in {SectionKey.COVER, SectionKey.TOC}:
                continue
            self._section_sheet(workbook, section, used)
        self._evidence_sheet(workbook, document, used)

        buffer = io.BytesIO()
        workbook.save(buffer)
        return RenderResult(
            payload=buffer.getvalue(), fmt=self.fmt,
            filename=self.filename(document, self.fmt),
            took_ms=round((time.perf_counter() - started) * 1000, 2),
        )

    # ------------------------------------------------------------------
    def _summary_sheet(self, workbook, document, used) -> None:
        sheet = workbook.create_sheet(_sheet_name("Summary", used))
        cover = document.cover

        sheet["A1"] = cover.company_name
        sheet["A1"].font = TITLE_FONT
        sheet["A2"] = f"{cover.ticker} · {cover.title}"
        sheet["A2"].font = MUTED_FONT

        row = 4
        for key, value in cover_pairs(document):
            sheet.cell(row, 1, key).font = LABEL_FONT
            number = _numeric(value)
            cell = sheet.cell(row, 2)
            if number is not None and key not in {"Ticker", "Report date"}:
                cell.value = number
                cell.number_format = (
                    "0.0%" if "%" in value else "#,##0.00" if "." in value
                    else "#,##0"
                )
            else:
                cell.value = _safe(value)
            cell.font = BODY_FONT
            row += 1

        if cover.data_warning:
            row += 1
            sheet.cell(row, 1, "Data quality").font = LABEL_FONT
            cell = sheet.cell(row, 2, cover.data_warning)
            cell.font = BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 2

        row += 1
        sheet.cell(row, 1, "Contents").font = Font(bold=True, size=11, color=NAVY)
        row += 1
        for section in document.ordered():
            if section.key in {SectionKey.COVER, SectionKey.TOC}:
                continue
            sheet.cell(row, 1, section.title).font = BODY_FONT
            status = "Insufficient evidence" if not section.sufficient else "Included"
            sheet.cell(row, 2, status).font = (
                MUTED_FONT if section.sufficient else Font(size=8, color="B45309")
            )
            row += 1

        row += 1
        sheet.cell(row, 1, "Statistics").font = Font(bold=True, size=11, color=NAVY)
        row += 1
        for key, value in document.statistics().items():
            sheet.cell(row, 1, key.replace("_", " ").title()).font = BODY_FONT
            sheet.cell(row, 2, value).font = BODY_FONT
            row += 1

        sheet.column_dimensions["A"].width = 34
        sheet.column_dimensions["B"].width = 62
        sheet.freeze_panes = "A4"

    def _section_sheet(self, workbook, section, used) -> None:
        sheet = workbook.create_sheet(_sheet_name(section.title, used))
        sheet["A1"] = section.title
        sheet["A1"].font = TITLE_FONT
        row = 3
        prose: list[str] = []
        chart_anchors: list[tuple[Block, int, int]] = []

        for block in section.blocks:
            kind = block.kind
            if kind is BlockKind.TABLE:
                start = row
                row = self._write_table(sheet, block, row)
                if block.caption or block.header:
                    chart_anchors.append((block, start, row))
                row += 1
            elif kind is BlockKind.KEY_VALUE:
                for key, value in block.pairs:
                    sheet.cell(row, 1, key).font = LABEL_FONT
                    number = _numeric(value)
                    cell = sheet.cell(row, 2)
                    cell.value = number if number is not None else _safe(value)
                    cell.font = BODY_FONT
                    row += 1
                row += 1
            elif kind is BlockKind.METRIC_GRID:
                for label, value, hint in block.metrics:
                    sheet.cell(row, 1, label).font = LABEL_FONT
                    number = _numeric(value)
                    cell = sheet.cell(row, 2)
                    cell.value = number if number is not None else _safe(value)
                    cell.font = BODY_FONT
                    sheet.cell(row, 3, hint).font = MUTED_FONT
                    row += 1
                row += 1
            elif kind is BlockKind.CHART:
                start = row
                row = self._write_chart_data(sheet, block, row)
                self._native_chart(sheet, block, start, row)
                row += 2
            elif kind in {BlockKind.PARAGRAPH, BlockKind.CALLOUT,
                          BlockKind.INSUFFICIENT, BlockKind.QUOTE}:
                prose.append(self._prose_text(block))
            elif kind is BlockKind.BULLETS:
                prose.extend(f"• {strip_markers(i)}" for i in block.items)
            elif kind is BlockKind.HEADING:
                sheet.cell(row, 1, block.text).font = Font(
                    bold=True, size=10, color=NAVY,
                )
                row += 1
            elif kind is BlockKind.CITATION_LIST:
                for entry in block.entries:
                    sheet.cell(row, 1, entry.key).font = LABEL_FONT
                    sheet.cell(row, 2, entry.render()).font = BODY_FONT
                    row += 1

        if prose:
            row += 1
            sheet.cell(row, 1, "Narrative").font = Font(
                bold=True, size=10, color=NAVY,
            )
            row += 1
            for text in prose:
                cell = sheet.cell(row, 1, _safe(text))
                cell.font = BODY_FONT
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                sheet.merge_cells(
                    start_row=row, start_column=1, end_row=row, end_column=8,
                )
                sheet.row_dimensions[row].height = max(
                    15, min(120, 13 * (len(text) // 110 + 1)),
                )
                row += 1

        sheet.column_dimensions["A"].width = 42
        for index in range(2, 12):
            sheet.column_dimensions[get_column_letter(index)].width = 16
        sheet.freeze_panes = "A3"

    @staticmethod
    def _prose_text(block: Block) -> str:
        if block.kind is BlockKind.CALLOUT:
            return f"{block.title}: {strip_markers(block.text)}"
        if block.kind is BlockKind.INSUFFICIENT:
            return block.text
        if block.kind is BlockKind.QUOTE:
            return f"\u201c{block.text}\u201d"
        return strip_markers(block.text)

    @staticmethod
    def _write_table(sheet: Worksheet, block, row: int) -> int:
        if block.caption:
            sheet.cell(row, 1, block.caption).font = Font(
                bold=True, size=10, color=NAVY,
            )
            row += 1
        if block.header:
            for index, heading in enumerate(block.header, start=1):
                cell = sheet.cell(row, index, heading)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            row += 1
        for row_index, values in enumerate(block.rows):
            for index, value in enumerate(values, start=1):
                cell = sheet.cell(row, index)
                number = _numeric(value)
                if number is not None and index > 1:
                    cell.value = number
                    cell.number_format = (
                        "0.0%" if "%" in value
                        else '#,##0.0"x"' if value.rstrip().endswith("x")
                        else "#,##0.00" if "." in value else "#,##0"
                    )
                else:
                    cell.value = _safe(value)
                cell.font = (
                    Font(size=9, bold=True) if row_index in block.emphasis_rows
                    else BODY_FONT
                )
                cell.border = BORDER
                if row_index % 2 == 1:
                    cell.fill = SURFACE_FILL
            row += 1
        return row

    @staticmethod
    def _write_chart_data(sheet: Worksheet, block, row: int) -> int:
        sheet.cell(row, 1, block.title).font = Font(bold=True, size=10, color=NAVY)
        row += 1
        header_row = row
        sheet.cell(row, 1, "Period").font = HEADER_FONT
        sheet.cell(row, 1).fill = HEADER_FILL
        for index, (name, _) in enumerate(block.series, start=2):
            cell = sheet.cell(row, index, name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        row += 1
        for position, label in enumerate(block.labels):
            sheet.cell(row, 1, _safe(label)).font = BODY_FONT
            for index, (_, values) in enumerate(block.series, start=2):
                value = values[position] if position < len(values) else None
                cell = sheet.cell(row, index)
                cell.value = value
                cell.font = BODY_FONT
                cell.number_format = "#,##0.00"
            row += 1
        return row

    @staticmethod
    def _native_chart(sheet: Worksheet, block, start: int, end: int) -> None:
        """Attach a live Excel chart to the data just written.

        Native rather than an embedded PNG: a spreadsheet chart that updates
        when the analyst changes an input is worth far more than a picture of
        one. Radar and heat-map kinds have no clean openpyxl equivalent, so
        they are left as data only.
        """
        if block.chart_kind in {ChartKind.SCORE_RADAR, ChartKind.SENSITIVITY,
                                ChartKind.PORTFOLIO_ALLOCATION}:
            return
        header_row = start + 1
        first_data = header_row + 1
        last_data = end - 1
        if last_data < first_data or not block.series:
            return

        chart = (
            LineChart() if block.chart_kind is ChartKind.MARGINS else BarChart()
        )
        chart.title = block.title
        chart.height = 7.5
        chart.width = 17
        chart.style = 2
        if block.y_unit:
            chart.y_axis.title = block.y_unit

        data = Reference(
            sheet, min_col=2, max_col=1 + len(block.series),
            min_row=header_row, max_row=last_data,
        )
        categories = Reference(
            sheet, min_col=1, min_row=first_data, max_row=last_data,
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        sheet.add_chart(chart, f"{get_column_letter(len(block.series) + 3)}{start}")

    def _evidence_sheet(self, workbook, document, used) -> None:
        sheet = workbook.create_sheet(_sheet_name("Evidence", used))
        sheet["A1"] = "Evidence and sources"
        sheet["A1"].font = TITLE_FONT
        sheet["A2"] = (
            "Every figure cited in this report, grouped by the engine that "
            "produced it."
        )
        sheet["A2"].font = MUTED_FONT

        row = 4
        for index, heading in enumerate(
            ["Reference", "Description", "Engine", "Value", "Unit", "Detail"],
            start=1,
        ):
            cell = sheet.cell(row, index, heading)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER
        row += 1

        for source, entries in sorted(
            evidence_by_source(document.evidence()).items(),
            key=lambda kv: kv[0].value,
        ):
            for entry in entries:
                sheet.cell(row, 1, entry.key).font = BODY_FONT
                sheet.cell(row, 2, entry.label).font = BODY_FONT
                sheet.cell(
                    row, 3, source.value.replace("_", " ").title()
                ).font = BODY_FONT
                cell = sheet.cell(row, 4)
                if isinstance(entry.value, (int, float)):
                    cell.value = entry.value
                    cell.number_format = "#,##0.00"
                else:
                    cell.value = _safe(str(entry.value)) if entry.value else "—"
                cell.font = BODY_FONT
                sheet.cell(row, 5, entry.unit).font = BODY_FONT
                sheet.cell(row, 6, entry.detail).font = MUTED_FONT
                for column in range(1, 7):
                    sheet.cell(row, column).border = BORDER
                row += 1

        widths = {"A": 26, "B": 40, "C": 20, "D": 16, "E": 10, "F": 40}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        sheet.freeze_panes = "A5"
        sheet.auto_filter.ref = f"A4:F{max(4, row - 1)}"
