"""Admin company management endpoints (Phase 2).

CRUD, bulk editor, CSV/Excel import & export, duplicate merging, logo upload,
and version history with rollback. All routes are permission-guarded
(``COMPANY_WRITE`` / ``COMPANY_READ``).
"""
from __future__ import annotations

import io

from fastapi import (
    APIRouter, Depends, File, HTTPException, Query, Request, UploadFile, status,
)
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.security import CurrentUser, get_current_user, require
from app.db.base import get_db
from app.domain.platform.identity import Permission
from app.schemas.company import (
    CompanyCreate, CompanyDetail, CompanyUpdate, CompanyVersionOut,
    ImportResult, MergeResult, PaginatedCompanies, CompanyBulkEditRequest,
    CompanyBulkEditResult,
)
from app.services.company_admin_service import (
    CompanyAdminError, CompanyAdminService,
)
from app.services.live_market import LiveMarketService

router = APIRouter(prefix="/admin/companies", tags=["admin-companies"])


def _service(db: Session = Depends(get_db)) -> CompanyAdminService:
    return CompanyAdminService(db)


def _detail(svc, company, *, attach_market: bool = False) -> CompanyDetail:
    detail = CompanyDetail.model_validate(company)
    if attach_market:
        return LiveMarketService.attach(detail, company, svc.db)
    return detail


# --------------------------------------------------------------------- CRUD
@router.get(
    "", response_model=PaginatedCompanies,
    summary="List companies (admin)",
    dependencies=[Depends(require(Permission.COMPANY_READ))],
)
def admin_list_companies(
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=200),
    search: str | None = None,
    sector: str | None = None,
    industry: str | None = None,
    exchange: str | None = None,
    listing_status: str | None = None,
    market_cap_min: float | None = None,
    market_cap_max: float | None = None,
    sort_by: str = Query("market_cap"),
    order: str = Query("desc"),
    include_deleted: bool = False,
    svc: CompanyAdminService = Depends(_service),
) -> PaginatedCompanies:
    total, rows = svc.list_companies(
        page=page, page_size=page_size, search=search, sector=sector,
        industry=industry, exchange=exchange, listing_status=listing_status,
        market_cap_min=market_cap_min, market_cap_max=market_cap_max,
        sort_by=sort_by, order=order, include_deleted=include_deleted,
    )
    return PaginatedCompanies(
        total=total, page=page, page_size=page_size,
        results=[_detail(svc, c) for c in rows],
    )


@router.get(
    "/filters", summary="Distinct sectors and industries",
    dependencies=[Depends(require(Permission.COMPANY_READ))],
)
def admin_filters(svc: CompanyAdminService = Depends(_service)) -> dict[str, list[str]]:
    return {"sectors": svc.sectors(), "industries": svc.industries()}


@router.post(
    "", response_model=CompanyDetail, status_code=status.HTTP_201_CREATED,
    summary="Add a company",
    dependencies=[Depends(require(Permission.COMPANY_WRITE))],
)
def admin_create_company(
    body: CompanyCreate,
    svc: CompanyAdminService = Depends(_service),
    user: CurrentUser = Depends(get_current_user),
) -> CompanyDetail:
    try:
        company = svc.create(
            body, actor_id=user.user_id, actor_email=user.email,
        )
    except CompanyAdminError as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(exc))
    svc.db.commit()
    return _detail(svc, company)


@router.get(
    "/{company_id}", response_model=CompanyDetail,
    summary="Get a company (admin)",
    dependencies=[Depends(require(Permission.COMPANY_READ))],
)
def admin_get_company(
    company_id: str, svc: CompanyAdminService = Depends(_service),
) -> CompanyDetail:
    company = svc.get(company_id, include_deleted=True)
    if company is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "company not found")
    return _detail(svc, company, attach_market=True)


@router.patch(
    "/{company_id}", response_model=CompanyDetail,
    summary="Edit a company",
    dependencies=[Depends(require(Permission.COMPANY_WRITE))],
)
def admin_update_company(
    company_id: str, body: CompanyUpdate,
    svc: CompanyAdminService = Depends(_service),
    user: CurrentUser = Depends(get_current_user),
) -> CompanyDetail:
    try:
        company = svc.update(
            company_id, body, actor_id=user.user_id, actor_email=user.email,
        )
    except CompanyAdminError as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(exc))
    svc.db.commit()
    return _detail(svc, company)


@router.delete(
    "/{company_id}", response_model=CompanyDetail,
    summary="Soft-delete a company (moves to recycle bin)",
    dependencies=[Depends(require(Permission.COMPANY_WRITE))],
)
def admin_soft_delete(
    company_id: str, svc: CompanyAdminService = Depends(_service),
    user: CurrentUser = Depends(get_current_user),
) -> CompanyDetail:
    try:
        company = svc.soft_delete(
            company_id, actor_id=user.user_id, actor_email=user.email,
        )
    except CompanyAdminError as exc:
        raise HTTPException(status.HTTP_404_NOT_FOUND, str(exc))
    svc.db.commit()
    return _detail(svc, company)


@router.post(
    "/{company_id}/restore", response_model=CompanyDetail,
    summary="Restore a soft-deleted company",
    dependencies=[Depends(require(Permission.COMPANY_WRITE))],
)
def admin_restore(
    company_id: str, svc: CompanyAdminService = Depends(_service),
    user: CurrentUser = Depends(get_current_user),
) -> CompanyDetail:
    try:
        company = svc.restore(
            company_id, actor_id=user.user_id, actor_email=user.email,
        )
    except CompanyAdminError as exc:
        raise HTTPException(status.HTTP_404_NOT_FOUND, str(exc))
    svc.db.commit()
    return _detail(svc, company)


@router.delete(
    "/{company_id}/permanent", status_code=status.HTTP_204_NO_CONTENT,
    summary="Permanently delete a company",
    dependencies=[Depends(require(Permission.COMPANY_WRITE))],
)
def admin_permanent_delete(
    company_id: str, svc: CompanyAdminService = Depends(_service),
    user: CurrentUser = Depends(get_current_user),
):
    try:
        svc.permanent_delete(
            company_id, actor_id=user.user_id, actor_email=user.email,
        )
    except CompanyAdminError as exc:
        raise HTTPException(status.HTTP_404_NOT_FOUND, str(exc))
    svc.db.commit()


# ------------------------------------------------------------ bulk + merge
@router.post(
    "/bulk-edit", response_model=CompanyBulkEditResult,
    summary="Apply spreadsheet-style bulk edits",
    dependencies=[Depends(require(Permission.COMPANY_WRITE))],
)
def admin_bulk_edit(
    body: CompanyBulkEditRequest,
    svc: CompanyAdminService = Depends(_service),
    user: CurrentUser = Depends(get_current_user),
) -> CompanyBulkEditResult:
    updated, created, errors = svc.bulk_edit(
        [i.model_dump(exclude_none=True) for i in body.items],
        actor_id=user.user_id, actor_email=user.email,
    )
    svc.db.commit()
    return CompanyBulkEditResult(updated=updated, created=created, errors=errors)


@router.post(
    "/merge", response_model=MergeResult,
    summary="Merge duplicate companies",
    dependencies=[Depends(require(Permission.COMPANY_WRITE))],
)
def admin_merge(
    keep_id: str, delete_ids: list[str] = Query(default=[]),
    svc: CompanyAdminService = Depends(_service),
    user: CurrentUser = Depends(get_current_user),
) -> MergeResult:
    try:
        keeper, merged = svc.merge_duplicates(
            keep_id=keep_id, delete_ids=delete_ids,
            actor_id=user.user_id, actor_email=user.email,
        )
    except CompanyAdminError as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(exc))
    svc.db.commit()
    return MergeResult(
        kept_id=keeper.id, kept_ticker=keeper.ticker,
        merged_ids=merged, removed_count=len(merged),
    )


# ------------------------------------------------------------ import/export
@router.post(
    "/import/csv", response_model=ImportResult,
    summary="Bulk import companies from CSV",
    dependencies=[Depends(require(Permission.COMPANY_WRITE))],
)
def admin_import_csv(
    file: UploadFile = File(...),
    svc: CompanyAdminService = Depends(_service),
    user: CurrentUser = Depends(get_current_user),
) -> ImportResult:
    import csv as _csv
    text = (file.file.read() or b"").decode("utf-8-sig")
    reader = _csv.DictReader(io.StringIO(text))
    rows = [dict(r) for r in reader]
    result = svc.import_rows(rows, actor_id=user.user_id, actor_email=user.email)
    svc.db.commit()
    return ImportResult(**result)


@router.post(
    "/import/xlsx", response_model=ImportResult,
    summary="Bulk import companies from Excel",
    dependencies=[Depends(require(Permission.COMPANY_WRITE))],
)
def admin_import_xlsx(
    file: UploadFile = File(...),
    svc: CompanyAdminService = Depends(_service),
    user: CurrentUser = Depends(get_current_user),
) -> ImportResult:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file.file.read()), read_only=True, data_only=True)
    ws = wb.active
    header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({header[i]: (row[i] if i < len(row) else None) for i in range(len(header))})
    result = svc.import_rows(rows, actor_id=user.user_id, actor_email=user.email)
    svc.db.commit()
    return ImportResult(**result)


@router.get(
    "/export/csv", summary="Export companies to CSV",
    dependencies=[Depends(require(Permission.COMPANY_READ))],
)
def admin_export_csv(
    include_deleted: bool = False,
    svc: CompanyAdminService = Depends(_service),
) -> StreamingResponse:
    import csv as _csv
    rows = svc.export_rows(include_deleted=include_deleted)
    buf = io.StringIO()
    fieldnames = list(rows[0].keys()) if rows else []
    writer = _csv.DictWriter(buf, fieldnames=fieldnames)
    writer.writeheader()
    for row in rows:
        writer.writerow({k: ("" if v is None else v) for k, v in row.items()})
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=companies.csv"},
    )


@router.get(
    "/export/xlsx", summary="Export companies to Excel",
    dependencies=[Depends(require(Permission.COMPANY_READ))],
)
def admin_export_xlsx(
    include_deleted: bool = False,
    svc: CompanyAdminService = Depends(_service),
) -> StreamingResponse:
    import openpyxl
    from openpyxl.utils import get_column_letter

    rows = svc.export_rows(include_deleted=include_deleted)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Companies"
    if rows:
        fieldnames = list(rows[0].keys())
        ws.append(fieldnames)
        for row in rows:
            ws.append([row.get(f) for f in fieldnames])
        for i, _name in enumerate(fieldnames, start=1):
            ws.column_dimensions[get_column_letter(i)].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=companies.xlsx"},
    )


# --------------------------------------------------------------- versions
@router.get(
    "/{company_id}/versions", response_model=list[CompanyVersionOut],
    summary="Company edit history",
    dependencies=[Depends(require(Permission.COMPANY_READ))],
)
def admin_company_versions(
    company_id: str, svc: CompanyAdminService = Depends(_service),
) -> list[CompanyVersionOut]:
    return svc.versions(company_id)


@router.post(
    "/{company_id}/rollback", response_model=CompanyDetail,
    summary="Roll a company back to an earlier version",
    dependencies=[Depends(require(Permission.COMPANY_WRITE))],
)
def admin_rollback(
    company_id: str, version: int = Query(...),
    svc: CompanyAdminService = Depends(_service),
    user: CurrentUser = Depends(get_current_user),
) -> CompanyDetail:
    try:
        company = svc.rollback(
            company_id, version, actor_id=user.user_id, actor_email=user.email,
        )
    except CompanyAdminError as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(exc))
    svc.db.commit()
    return _detail(svc, company)


# ---------------------------------------------------------------- logo upload
@router.post(
    "/{company_id}/logo", response_model=CompanyDetail,
    summary="Set a company logo / favicon URL",
    dependencies=[Depends(require(Permission.COMPANY_WRITE))],
)
def admin_upload_logo(
    company_id: str, logo_url: str | None = Query(None),
    favicon_url: str | None = Query(None),
    svc: CompanyAdminService = Depends(_service),
    user: CurrentUser = Depends(get_current_user),
) -> CompanyDetail:
    try:
        company = svc.update(
            company_id,
            CompanyUpdate(logo_url=logo_url, favicon_url=favicon_url),
            actor_id=user.user_id, actor_email=user.email,
        )
    except CompanyAdminError as exc:
        raise HTTPException(status.HTTP_404_NOT_FOUND, str(exc))
    svc.db.commit()
    return _detail(svc, company)
