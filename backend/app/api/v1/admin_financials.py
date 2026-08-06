"""Admin financial-statements endpoints (Phase 3).

Editable income statement, balance sheet, cash flow, ratios, quarterly results,
shareholding pattern and corporate actions; bulk CSV/Excel/JSON import; annual
report PDF import (extract → preview → approve); and fact-level version history
with rollback. All routes are permission-guarded.
"""
from __future__ import annotations

import io
import json

from fastapi import (
    APIRouter, Depends, File, HTTPException, Query, Request, UploadFile, status,
)
from sqlalchemy.orm import Session

from app.core.security import CurrentUser, get_current_user, require
from app.db.base import get_db
from app.domain.platform.identity import Permission
from app.schemas.financials import (
    FinancialBulkResult, FinancialStatementsOut, FinancialVersionOut,
    QuarterlyResultIn, ShareholdingIn, CorporateActionIn, CorporateActionUpdate,
)
from app.services.company_admin_service import CompanyAdminError
from app.services.financial_admin_service import (
    FinancialAdminError, FinancialAdminService,
)

router = APIRouter(prefix="/admin/financials", tags=["admin-financials"])


def _service(db: Session = Depends(get_db)) -> FinancialAdminService:
    return FinancialAdminService(db)


def _read_or_404(exc: FinancialAdminError) -> HTTPException:
    return HTTPException(status.HTTP_404_NOT_FOUND, str(exc))


# ------------------------------------------------------------- read
@router.get(
    "/{company_id}/statements", summary="Annual statements + ratios",
    dependencies=[Depends(require(Permission.COMPANY_READ))],
)
def get_statements(company_id: str, svc: FinancialAdminService = Depends(_service)):
    try:
        return svc.statements(company_id)
    except FinancialAdminError as exc:
        raise _read_or_404(exc)


@router.get(
    "/{company_id}/quarterly", summary="Quarterly results",
    dependencies=[Depends(require(Permission.COMPANY_READ))],
)
def get_quarterly(company_id: str, svc: FinancialAdminService = Depends(_service)):
    try:
        return {"items": svc.quarterly(company_id)}
    except FinancialAdminError as exc:
        raise _read_or_404(exc)


@router.get(
    "/{company_id}/shareholding", summary="Shareholding pattern",
    dependencies=[Depends(require(Permission.COMPANY_READ))],
)
def get_shareholding(company_id: str, svc: FinancialAdminService = Depends(_service)):
    try:
        return {"items": svc.shareholding(company_id)}
    except FinancialAdminError as exc:
        raise _read_or_404(exc)


@router.get(
    "/{company_id}/corporate-actions", summary="Corporate actions",
    dependencies=[Depends(require(Permission.COMPANY_READ))],
)
def get_corporate_actions(company_id: str, svc: FinancialAdminService = Depends(_service)):
    try:
        return {"items": svc.corporate_actions(company_id)}
    except FinancialAdminError as exc:
        raise _read_or_404(exc)


# ------------------------------------------------------------- annual facts
@router.put(
    "/{company_id}/facts", response_model=FinancialBulkResult,
    summary="Bulk upsert annual financial facts",
    dependencies=[Depends(require(Permission.COMPANY_WRITE))],
)
def put_facts(company_id: str, facts: list[dict],
              svc: FinancialAdminService = Depends(_service),
              user: CurrentUser = Depends(get_current_user)):
    try:
        result = svc.upsert_facts(company_id, facts, actor_id=user.user_id, actor_email=user.email)
    except FinancialAdminError as exc:
        raise _read_or_404(exc)
    svc.db.commit()
    return result


# ------------------------------------------------------------- quarterly
@router.put(
    "/{company_id}/quarterly", response_model=FinancialBulkResult,
    summary="Bulk upsert quarterly results",
    dependencies=[Depends(require(Permission.COMPANY_WRITE))],
)
def put_quarterly(company_id: str, items: list[QuarterlyResultIn],
                  svc: FinancialAdminService = Depends(_service),
                  user: CurrentUser = Depends(get_current_user)):
    try:
        result = svc.upsert_quarterly(
            company_id, [i.model_dump(exclude_none=True) for i in items],
            actor_id=user.user_id, actor_email=user.email,
        )
    except FinancialAdminError as exc:
        raise _read_or_404(exc)
    svc.db.commit()
    return result


@router.delete(
    "/{company_id}/quarterly/{year}/{quarter}", status_code=status.HTTP_204_NO_CONTENT,
    summary="Delete a quarterly result",
    dependencies=[Depends(require(Permission.COMPANY_WRITE))],
)
def delete_quarterly(company_id: str, year: int, quarter: int,
                     svc: FinancialAdminService = Depends(_service),
                     user: CurrentUser = Depends(get_current_user)):
    try:
        svc.delete_quarterly(company_id, year, quarter, actor_id=user.user_id, actor_email=user.email)
    except FinancialAdminError as exc:
        raise _read_or_404(exc)
    svc.db.commit()


# ------------------------------------------------------------- shareholding
@router.put(
    "/{company_id}/shareholding", response_model=FinancialBulkResult,
    summary="Bulk upsert shareholding pattern",
    dependencies=[Depends(require(Permission.COMPANY_WRITE))],
)
def put_shareholding(company_id: str, items: list[ShareholdingIn],
                     svc: FinancialAdminService = Depends(_service),
                     user: CurrentUser = Depends(get_current_user)):
    try:
        result = svc.upsert_shareholding(
            company_id, [i.model_dump(exclude_none=True) for i in items],
            actor_id=user.user_id, actor_email=user.email,
        )
    except FinancialAdminError as exc:
        raise _read_or_404(exc)
    svc.db.commit()
    return result


# ------------------------------------------------------------- corporate actions
@router.post(
    "/{company_id}/corporate-actions", summary="Add a corporate action",
    dependencies=[Depends(require(Permission.COMPANY_WRITE))],
)
def add_action(company_id: str, body: CorporateActionIn,
               svc: FinancialAdminService = Depends(_service),
               user: CurrentUser = Depends(get_current_user)):
    try:
        action = svc.add_corporate_action(
            company_id, body.model_dump(exclude_none=True),
            actor_id=user.user_id, actor_email=user.email,
        )
    except FinancialAdminError as exc:
        raise _read_or_404(exc)
    svc.db.commit()
    return {c.name: getattr(action, c.name) for c in action.__table__.c}


@router.patch(
    "/{company_id}/corporate-actions/{action_id}", summary="Update a corporate action",
    dependencies=[Depends(require(Permission.COMPANY_WRITE))],
)
def update_action(company_id: str, action_id: int, body: CorporateActionUpdate,
                  svc: FinancialAdminService = Depends(_service),
                  user: CurrentUser = Depends(get_current_user)):
    try:
        action = svc.update_corporate_action(
            company_id, action_id, body.model_dump(exclude_none=True),
            actor_id=user.user_id, actor_email=user.email,
        )
    except FinancialAdminError as exc:
        raise _read_or_404(exc)
    svc.db.commit()
    return {c.name: getattr(action, c.name) for c in action.__table__.c}


@router.delete(
    "/{company_id}/corporate-actions/{action_id}", status_code=status.HTTP_204_NO_CONTENT,
    summary="Delete a corporate action",
    dependencies=[Depends(require(Permission.COMPANY_WRITE))],
)
def delete_action(company_id: str, action_id: int,
                  svc: FinancialAdminService = Depends(_service),
                  user: CurrentUser = Depends(get_current_user)):
    try:
        svc.delete_corporate_action(company_id, action_id, actor_id=user.user_id, actor_email=user.email)
    except FinancialAdminError as exc:
        raise _read_or_404(exc)
    svc.db.commit()


# ------------------------------------------------------------- bulk import
@router.post(
    "/{company_id}/bulk-import", response_model=FinancialBulkResult,
    summary="Bulk import financials from CSV / Excel / JSON",
    dependencies=[Depends(require(Permission.COMPANY_WRITE))],
)
def bulk_import(
    company_id: str, file: UploadFile = File(...), kind: str = Query("facts"),
    svc: FinancialAdminService = Depends(_service),
    user: CurrentUser = Depends(get_current_user),
):
    """`kind` in {facts, quarterly, shareholding}. Detects CSV/Excel/JSON by
    filename extension and dispatches to the matching bulk upsert."""
    import csv as _csv
    filename = (file.filename or "").lower()
    data = file.file.read()
    try:
        if filename.endswith(".json"):
            rows = json.loads(data.decode("utf-8-sig"))
            rows = rows if isinstance(rows, list) else rows.get("items", [])
        elif filename.endswith((".xlsx", ".xlsm")):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({header[i]: (row[i] if i < len(row) else None) for i in range(len(header))})
        else:
            text = data.decode("utf-8-sig")
            reader = _csv.DictReader(io.StringIO(text))
            rows = [dict(r) for r in reader]
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"could not parse file: {exc}")

    def _num(v):
        if v in (None, ""):
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return v

    if kind == "facts":
        normalized = [
            {"fiscal_year": int(_num(r.get("fiscal_year") or r.get("year"))),
             "line_item": str(r.get("line_item") or r.get("item") or "").strip(),
             "value": _num(r.get("value"))}
            for r in rows if r.get("fiscal_year") or r.get("year")
        ]
        result = svc.upsert_facts(company_id, normalized, actor_id=user.user_id, actor_email=user.email)
    elif kind == "quarterly":
        normalized = [
            {"fiscal_year": int(_num(r.get("fiscal_year") or r.get("year"))),
             "quarter": int(_num(r.get("quarter"))),
             **{k: _num(v) for k, v in r.items() if k not in ("fiscal_year", "year", "quarter")}}
            for r in rows if (r.get("fiscal_year") or r.get("year")) and r.get("quarter")
        ]
        result = svc.upsert_quarterly(company_id, normalized, actor_id=user.user_id, actor_email=user.email)
    elif kind == "shareholding":
        normalized = [
            {"fiscal_year": int(_num(r.get("fiscal_year") or r.get("year"))),
             "quarter": int(_num(r.get("quarter"))),
             **{k: _num(v) for k, v in r.items() if k not in ("fiscal_year", "year", "quarter")}}
            for r in rows if (r.get("fiscal_year") or r.get("year")) and r.get("quarter")
        ]
        result = svc.upsert_shareholding(company_id, normalized, actor_id=user.user_id, actor_email=user.email)
    else:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"unknown kind '{kind}'")
    svc.db.commit()
    return result


# ------------------------------------------------------------- versions
@router.get(
    "/{company_id}/versions", response_model=list[FinancialVersionOut],
    summary="Financial version history",
    dependencies=[Depends(require(Permission.COMPANY_READ))],
)
def get_versions(company_id: str, svc: FinancialAdminService = Depends(_service)):
    try:
        return svc.versions(company_id)
    except FinancialAdminError as exc:
        raise _read_or_404(exc)


@router.post(
    "/{company_id}/rollback", summary="Roll financials back to a version",
    dependencies=[Depends(require(Permission.COMPANY_WRITE))],
)
def rollback(company_id: str, version: int = Query(...),
             svc: FinancialAdminService = Depends(_service),
             user: CurrentUser = Depends(get_current_user)):
    try:
        svc.rollback(company_id, version, actor_id=user.user_id, actor_email=user.email)
    except (FinancialAdminError, CompanyAdminError) as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(exc))
    svc.db.commit()
    return {"status": "ok", "rolled_back_to": version}
